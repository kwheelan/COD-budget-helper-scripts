"""
This script will automatitcally create the well-formatted HARI report from the emailed report.

K. Wheelan 2025
"""

# ======================================================================================
# SECION 1: Package imports
# 
# Do not edit this section.
# ======================================================================================

import os # operating system package for opening and saving files
import pandas as pd # pandas does data manipulation
from openpyxl import load_workbook # openpyxl has Excel functions
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.formatting.rule import CellIsRule

# ======================================================================================
# SECION 2: Set filepaths
# 
# This is where file paths are set. Change some of these variables to match the right
# file locations on your personal computer. See the comments below.
# ======================================================================================

# Change the right side of this equality to be the filepath of the original HARI file
# (ie. the unedited file received via email). Make sure it is surrounded by quotes and 
# has the leading r (the r tells it to ignore the \s).
INPUT = r'input_data\HARI\FA_OBJParent_V2_RP_Hari_FY25_2230Daily.xlsx'

# Change the right side of this equality to be the folder where you want the converted
# HARI to be saved. Make sure it is surrounded by quotes and has the leading r.
OUTPUT = r'output/HARI/'

# Edit the final file name here if you like. Keep the quotes.
new_file_name = 'converted_HARI.xlsx' 
output_file = os.path.join(OUTPUT, new_file_name)

# ======================================================================================
# SECTION 3: Define columns 
# 
# This is where we will define steps for the conversion. Edit if needed.
# ======================================================================================

# These are columns you want to delete from the original HARI. To add to the list, 
# add a comma after the last element but before the closing ], hit enter, and add the
# new column name exactly as in the Excel doc but inside quotes.
cols_to_delete = ['FAC_Numbers',
                'FAC_Names',
                'PARENT2_VALUE',
                'PARENT2_DESCRIPTION',
                'PARENT4_VALUE',
                'PARENT4_DESCRIPTION',
                'JournalStatus',
                'FundsReservedStatus']

# These are the new column names for the converted HARI report, listed in order. If 
# you adjust the list, just make sure every element is split with a comma and all
# column names are inside quotes.
new_col_names = ['Fund #', 'Fund Name',
           'Appropriation #', 'Appropriation Name',
           'Cost Center #', 'Cost Center Name',
           'Account Type',
           'Object #', 'Object Name',
           'Project #', 'Project Name',
           'Activity #', 'Activity Name',
           'Interfund #', 'Interfund Name',
           'Adopted Budget', 'Amended Budget', 
           'Commitment', 'Obligation', 'Actual', 
           'Funds Available'
           ]

# This the list of columns in the desired order for the final HARI report. Changing
# the list order here will change the column order in the final Excel document. Make 
# sure that every column listed in new_col_names defined above is also in this list.
# Newly created columns (Future #, Future Name, Account String, and Dept #) should 
# also be in this list. 
new_column_order = ['Account Type', 'Dept #',
           'Fund #', 'Fund Name',
           'Appropriation #', 'Appropriation Name',
           'Cost Center #', 'Cost Center Name',
           'Object #', 'Object Name',
           'Project #', 'Project Name',
           'Activity #', 'Activity Name',
           'Interfund #', 'Interfund Name',
           'Future #', 'Future Name',
           'Account String',
           'Adopted Budget', 'Amended Budget', 
           'Commitment', 'Obligation', 'Actual', 'Funds Available'
           ]

# ======================================================================================
# SECTION 4: Data manipulation 
# 
# This section will read in the data from the original HARI report, delete desired 
# columns, rename columns, create new columns, reorder columns, aggregate across each 
# unique account string, and then delete rows of all zeroes. You shouldn't have to 
# edit this section.
# ======================================================================================

# Read in data from original Excel sheet
df = pd.read_excel(INPUT, sheet_name='Sheet4', 
                   dtype=str,
                   # nrows=100,
                   skiprows=9,
                   )

# Delete specified columns
df.drop(cols_to_delete, axis=1, inplace=True)

# Rename headers with new column names
df.columns = new_col_names

# Add new columns and fill with desired default values
df['Future #'] = '000000'
df['Future Name'] = 'Undefined Future'
# acount string is fund-appropriation-CC-object-project-activity-interfund-future
df['Account String'] = df.apply(lambda row: f"{row['Fund #']}-{row['Appropriation #']}-{row['Cost Center #']}-{row['Object #']}-{row['Project #']}-{row['Activity #']}-{row['Interfund #']}-{row['Future #']}", axis=1)
# department number is the first 2 digits of the cost center
df['Dept #'] =  df['Cost Center #'].str[:2]

# Adjust column order
df = df[new_column_order]

# Data "compressed" to put repeat account strings on a single line 
# (ie aggregated across each unique account string)
group_columns = new_column_order[0:19]
value_cols = new_column_order[19:]
df[value_cols] = df[value_cols].astype(float)
df = df.groupby(group_columns, as_index=False).sum()

# Rows with $0.00 in columns T thru Y deleted
df = df.loc[~(df[value_cols].sum(axis=1) == 0)]
# Reset the index of the DataFrame
df.reset_index(drop=True, inplace=True)

# ======================================================================================
# SECTION 5: Create and save Excel file 
# 
# This section will create the final excel file and save to your previously set file 
# location. You should not have to edit anything here.
# ======================================================================================

# Save DataFrame to an Excel file with table formatting
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Write DataFrame to Excel, starting from the first row
    df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=2)
    worksheet = writer.sheets['Sheet1']

    # Get the dimensions of the DataFrame
    rows, cols = df.shape

    # Adjust column widths to fit the header and contents
    for col in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        # add extra width for table arrows
        worksheet.column_dimensions[col[0].column_letter].width = max_length + 6

    # Define the range for the table
    table_range = f"A3:{chr(65 + cols - 1)}{rows + 3}"

    # Create and style the table (with a style that has gray headers)
    table = Table(displayName="Table1", ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium2",  # Choose a style that has gray headers
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,  # Set to True if you want striped rows
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    worksheet.add_table(table)

    # Set the header color to gray
    gray_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for cell in worksheet["3:3"]:
        cell.fill = gray_fill

    # Group name columns for a collapsible section
    for col in 'DFHJLNPR':
        worksheet.column_dimensions.group(col, col, hidden=False)

    # Total & Subtotal formulas above columns T thru Y
    last_row_number = rows + 4  # accounting for two inserted rows and one header row
    # Add labels
    worksheet['S1'] = 'Total:'
    worksheet['S2'] = 'Subtotal:'
    # Right-align every cell
    worksheet['S1'].alignment = Alignment(horizontal='right')
    worksheet['S2'].alignment = Alignment(horizontal='right')
    # define the formulas for the summary rows at the top of the file
    for idx, col_letter in enumerate('tuvwxy', start=1):  
        worksheet[f"{col_letter}{1}"] = f"=SUBTOTAL(9,{col_letter}4:{col_letter}{last_row_number - 1})"
        worksheet[f"{col_letter}{2}"] = f"=SUM({col_letter}4:{col_letter}{last_row_number - 1})"

    # Apply right alignment and currency format to specified columns
    currency_columns = ['Adopted Budget', 'Amended Budget', 
           'Commitment', 'Obligation', 'Actual', 'Funds Available']
    
    # Add conditional formatting to make negatives red
    red_text_rule = CellIsRule(operator='lessThan', formula=['0'], font=Font(color="FF0000"))

    for col_idx in range(1, cols + 1):
        column_letter = chr(64 + col_idx)
        worksheet.conditional_formatting.add(f"{column_letter}1:{column_letter}{rows + 3}", red_text_rule)

    for col in worksheet.iter_cols(min_col=1, max_col=cols, min_row=1):
        header = col[2].value
        for cell in col[:2] + col[3:]:  # Skip the header cell
            # Right-align cells
            cell.alignment = Alignment(horizontal='right')
            # Apply currency format to specified columns
            if header in currency_columns:
                cell.number_format = cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
                