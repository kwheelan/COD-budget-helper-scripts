"""
K. Wheelan
Updated Feb 2025

Iterate through each file in detail_sheets,
 copy to the master DS file and save it
"""

# ======================================================================================
# SECION 1: Package imports
# 
# Do not edit this section.
# ======================================================================================

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import os
import shutil
# import custom functions
from utils.excel_utils import copy_cols, last_data_row, col_range

# ======================================================================================
# SECION 2: Set filepaths
# 
# This is where file paths are set. Change some of these variables to match the right
# file locations on your personal computer. See the comments below.
# ======================================================================================

# Change the right side of this equality to be the filepath of the folder where 
# the detail sheets to be combines are saved. Make sure it is surrounded by quotes and 
# has the leading r (the r tells it to ignore the \s).DATA = 'input_data'
SOURCE_FOLDER = r'C:/Users/katrina.wheelan/OneDrive - City of Detroit/Documents - M365-OCFO-Budget/BPA Team/FY 2026/1. Budget Development/08A. Deputy Budget Director Recommend/'

# Change the right side of this equality to be the filepath of the folder where
# you want the final master detail sheet to be saved. Make sure it is surrounded by quotes and 
# has the leading r (the r tells it to ignore the \s).DATA = 'input_data'
OUTPUT = r'output/master_DS/'

# Change the right side of this equality to be the name of the final file (master 
# detail sheet). Make sure it's surrounded by quotes.
final_file_name = 'master_detail_sheet_FY26.xlsx'

# Make sure this is set to the location of your detail sheet template
template_file = r'../files/FY26 Detail Sheet Template.xlsx'

# Do not edit the line below
dest_file = os.path.join(OUTPUT, final_file_name)

# ======================================================================================
# SECION 3: Set columns for copying
# 
# Edit this section if additional columns are added to the detail sheet (ie. extra 
# approval columns)
# ======================================================================================

# For each tab, this include the column range to copy. For example, col_range('A', 'BE') 
# will copy columns A through BE. To include additional columns, replace 'BE' with another 
# column. 
SHEETS = {
    'FTE, Salary-Wage, & Benefits' : { 
        # Below are the columns to copy from each DS in the FTE tab
        'cols' : col_range('A', 'BE'), 
        'start_row' : 15 },
    'Overtime & Other Personnel' : {
        # Below are the columns to copy from each DS in the OT tab
        'cols' : col_range('A', 'AN'),
        'start_row' : 15 },
    'Non-Personnel' : {
        # Below are the columns to copy from each DS in the NP tab
        'cols' : col_range('A', 'AH'),
        'start_row' : 19 },
    'Revenue' : {
        # Below are the columns to copy from each Revenue in the FTE tab
        'cols' : col_range('A', 'AE'),
        'start_row' : 15 }
}

# ======================================================================================
# SECION 4: Set columns for the summary tab
# 
# Optionally adjust this section to reflect the correct columns for summarization
# ======================================================================================

# Adjust as needed to make sure the correct columns are recorded.
SUMMARY_SECTIONS = {
    'FTE' : {
        # This the exact name of the tab with FTE information
        'tab' : 'FTE, Salary-Wage, & Benefits',
        # This is the column that counts the number of FTE
        'count_col' : 'AZ',
        # this is the column that says whether a line item is baseline or supplemental
        'baseline_col' : 'L'
    }, 
    'Salary & Benefits' : {
        # This the exact name of the tab with FTE information
        'tab' : 'FTE, Salary-Wage, & Benefits',
        # This is the column that has the SALARY amount for each line item
        'count_col': 'BC',
        # this is the column that says whether a line item is baseline or supplemental
        'baseline_col' : 'L'
    }, 
    'Non-Personnel' : {
        # This the exact name of the tab with non-personel information
        'tab' : 'Non-Personnel',
        # This is the column that has the amount for each NON-PERSONNEL line item
        'count_col': 'AG',
        # this is the column that says whether a line item is baseline or supplemental
        'baseline_col' : 'O'
    }, 
    'Overtime' : {
        # This the exact name of the tab with OT information
        'tab' : 'Overtime & Other Personnel',
        # This is the column that has the amount for each OVERTIME line item
        'count_col': 'AM',
        # this is the column that says whether a line item is baseline or supplemental
        'baseline_col' : 'N'
    }, 
    'Revenue' : {
        # This the exact name of the tab with revenue information
        'tab' : 'Revenue',
        # This is the column that has the amount for each REVENUE line item
        'count_col': 'AD',
        # this is the column that says whether a line item is baseline or supplemental
        'baseline_col' : 'P'
    }}

# ======================================================================================
# SECION 5: Departments to include in master detail sheet
# 
# ======================================================================================

# Specify departments to include in the master detail sheet. Add any additional departments
# with quotes around the name, and remove departments from the list by deleting them or adding
# a pound sign (#) before the first quote on the department's line. Departments with # (might look 
# green depending on your display settings) will be excluded.
INCLUDE = [
    'Airport',
    'BSEED',
    '16 CDD',
    '18 DSLP',
    '19 DPW',
    '20 DDoT',
   # '23 OCFO',
    '24 Fire',
    '25 Health',
    '28 HR',
    '29 CRIO'
    '31 DoIT',
    '32 Law',
    '33 Mayor',
    '34 Parking',
    '35 Non-Dept',
    '36 HRD Classic',
    '36 HRD JET Team',
    '38 PLD',
    '43 PDD',
    '45 DAH',
    '47 GSD',
    '50 OAG',
    '51 BZA',
    '52 Council',
    '53 Ombuds',
    '54 OIG',
    '60 36',
    '70 City Clerk',
    '71 Elections'
   # '72 Library'
]

# ======================================================================================
# SECION 6: Functions and class definitions
# 
# DO NOT EDIT BELOW THIS LINE
# ======================================================================================

# Set the gray fill for background color on Excel cells
gray_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

def move_data(detail_sheet, destination_file):
    # Load the workbooks
    #source_wb_values = load_workbook(f'{source_folder}/{detail_sheet}', data_only=True)
    source_wb_formulas = load_workbook(detail_sheet, data_only=False)
    destination_wb = load_workbook(destination_file, data_only=False)

    for sheet in SHEETS:
        if sheet in source_wb_formulas.sheetnames:
            source_ws_formulas = source_wb_formulas[sheet]
            dest_ws = destination_wb[sheet]
            config = SHEETS[sheet]

            # Determine the starting row in the destination sheet for pasting data
            dest_start_row = last_data_row(dest_ws, config['start_row']) + 1
            source_row_end = last_data_row(source_ws_formulas, n_header_rows=config['start_row']) + 1
            
            # Copy formula columns
            copy_cols(source_ws_formulas, dest_ws, config['cols'], 
                    source_row_start=config['start_row'], 
                    destination_row_start=dest_start_row,
                    source_row_end=source_row_end)
            
            name = detail_sheet.split('\\')[-1]
            print(f'Copied {sheet} data from {name}.')

    destination_wb.save(destination_file)

def create_summary(destination_file):

    fund_col = 'D'
    approp_col = 'G'

    # load workbook and formulas
    destination_wb = load_workbook(destination_file, data_only=False)
    # placeholder for funds
    funds = {}

    # go through each sheet and collect all unique funds
    for sheet in SHEETS:
        fund_data = destination_wb[sheet][fund_col]
        approp_data = destination_wb[sheet][approp_col]
        # start serach after header
        for i in range(SHEETS[sheet]['start_row'], len(fund_data)):
            fund = fund_data[i].value
            approp = approp_data[i].value
            if fund and (fund not in funds):
                funds[fund] = []
            if approp and (approp not in funds[fund]) and str(approp) not in funds[fund]:
                funds[fund].append(approp)

    # Transform the dictionary into a 2-column list
    rows = []
    for key, values in funds.items():
        for value in values:
            rows.append([key, value])
        rows.append([key, 'Subtotal'])

    # Copy to the summary tab
    summary = destination_wb['Summary']

    def summary_formula(section, baseOrSupp, fund_cell, approp_cell):
        """ Creating the SUMIFS formulas for summary tab """
        # get sheet for relevant data
        tab = SUMMARY_SECTIONS[section]['tab']
        # column to aggregate in that sheet
        col = SUMMARY_SECTIONS[section]['count_col']
        # baseline/supplemental column
        b_s_col = SUMMARY_SECTIONS[section]['baseline_col']
        
        if approp_cell.value == 'Subtotal':
            return f'=SUMIFS(\'{tab}\'!${col}:${col}, \'{tab}\'!${b_s_col}:${b_s_col}, "{baseOrSupp}", \'{tab}\'!${fund_col}:${fund_col}, {fund_cell.coordinate})'
        return f'=SUMIFS(\'{tab}\'!${col}:${col}, \'{tab}\'!${b_s_col}:${b_s_col}, "{baseOrSupp}", \'{tab}\'!${fund_col}:${fund_col}, {fund_cell.coordinate}, \'{tab}\'!${approp_col}:${approp_col}, {approp_cell.coordinate})'

    header_rows = 2
    # Write the transformed data to the sheet
    for row_ix in range(len(rows)):
        active_row = row_ix+(header_rows+1)
        # for each row (ie. fund/approp combo)
        fund_cell = summary.cell(row=active_row, column=1)
        approp_cell = summary.cell(row=active_row, column=2)
        fund_cell.value, approp_cell.value = rows[row_ix]

        # for each section: FTE, Salary and Benefits, NP, OT, Revenue
        for section_ix in range(len(SUMMARY_SECTIONS)):
            # get section name
            section = list(SUMMARY_SECTIONS.keys())[section_ix]
            # fetch cell locations and fill with formulas
            baseline_cell  = summary.cell(row=active_row, column=3+(section_ix*3))
            baseline_cell.value = summary_formula(section, 'Baseline', fund_cell, approp_cell)
            supplemental_cell = summary.cell(row=active_row, column=4+(section_ix*3))
            supplemental_cell.value = summary_formula(section, 'Supplemental', fund_cell, approp_cell)
            total_cell = summary.cell(row=active_row, column=5+(section_ix*3))
            total_cell.value = f'={baseline_cell.coordinate} + {supplemental_cell.coordinate}'

        # add total column to sum all subtotals
        total_col_cell = summary.cell(row=active_row, column=18)
        total_col_cell.value = f'=SUM(H{active_row}, K{active_row}, N{active_row})'

        # format as USD
        for col in range(6,19):
            cell = summary.cell(row=active_row, column=col)
            cell.number_format = cell.number_format = '"$"#,##0' 

        # REPLACED WITH CONDITIONAL FORMATTING IN WB
        # # bold row if total row
        # if approp_cell.value == 'Subtotal':
        #     for col in range(1, 19):
        #         cell = summary.cell(row=active_row, column=col)
        #         # turn row gray with bold
        #         cell.font = Font(bold=True)
        #         cell.fill = gray_fill

    # Save the workbook
    destination_wb.save(filename=destination_file)

# MOVE TO UTILS

def find_DS(folder, verbose = False):
    # Get full file path
    folder_fp = os.path.join(SOURCE_FOLDER, folder)

    # grab list of all files in that folder
    files = os.listdir(folder_fp)

    # generate list of reviewed detail sheets
    reviewed_DS = []
    for file in files:
        if ('Budget Director' in file) and '.xlsx' in file:
            reviewed_DS.append(os.path.join(SOURCE_FOLDER, folder, file))

    # return message
    if len(reviewed_DS) > 1:
        message = f'Multiple potential reviewed detail sheets: {reviewed_DS}'
    elif len(reviewed_DS) == 0:
        message = f'No reviewed detail sheet found in {folder}'
    if verbose:
        print(message)

    return reviewed_DS

def create_file_list(verbose = False):
    # all folders in analyst review section
    DS_FOLDERS = os.listdir(SOURCE_FOLDER)

    # initialize list and add viable files to it
    DS_list = []
    for folder in DS_FOLDERS:
        folder_fp = os.path.join(SOURCE_FOLDER, folder)
        if(os.path.isdir(folder_fp)):
            DS_list += find_DS(folder, verbose)

    return DS_list

def include_dept(file):
    for dept in INCLUDE:
        if dept in file:
            return dept
    return False

def exclude_dept(file):
    for dept in EXCLUDE:
        if dept in file:
            return dept
    return False



#================== Program on run ============================================

def main():
    # copy the template
    shutil.copy(template_file, dest_file)

    # get list of DS files
    # DS_list = [file for file in create_file_list() if not exclude_dept(file)]
    DS_list = [file for file in create_file_list() if include_dept(file)]
    for detail_sheet in DS_list:
        move_data(detail_sheet, dest_file)
    create_summary(dest_file)
    print("Created summary tab")

if __name__ == '__main__':
    main()
