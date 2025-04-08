# app.py

from flask import Flask, request, render_template, send_file
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Alignment, numbers, Font
from openpyxl.formatting.rule import CellIsRule
import os

app = Flask(__name__)

# Create a folder to store upload files and output if it doesn't exist
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'output'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Check if the post request has the file part
        if 'file' not in request.files:
            return 'No file part'

        file = request.files['file']
        if file.filename == '':
            return 'No selected file'

        if file:
            input_filepath = os.path.join(UPLOAD_FOLDER, file.filename)
            file.save(input_filepath)
            print('saved input file')

            # Process the file
            output_filepath = process_file(input_filepath, file.filename)

            # Provide a link to download the result
            return f'File processed successfully. <a href="/download/{os.path.basename(output_filepath)}">Download</a>'

    return render_template('index.html')

@app.route('/download/<filename>')
def download_file(filename):
    return send_file(os.path.join(OUTPUT_FOLDER, filename), as_attachment=True)

def process_file(filepath, original_filename):
    # Read in data
    df = pd.read_excel(filepath, sheet_name='Sheet4', 
                       dtype=str,
                       skiprows=9)

    # Delete columns
    to_delete = ['FAC_Numbers', 'FAC_Names', 'PARENT2_VALUE', 
                 'PARENT2_DESCRIPTION', 'PARENT4_VALUE', 
                 'PARENT4_DESCRIPTION', 'JournalStatus', 
                 'FundsReservedStatus']
    df.drop(to_delete, axis=1, inplace=True)

    # Rename headers
    headers = ['Fund #', 'Fund Name', 'Appropriation #', 
               'Appropriation Name', 'Cost Center #', 
               'Cost Center Name', 'Account Type', 'Object #', 
               'Object Name', 'Project #', 'Project Name', 
               'Activity #', 'Activity Name', 'Interfund #', 
               'Interfund Name', 'Adopted Budget', 
               'Amended Budget', 'Commitment', 'Obligation', 
               'Actual', 'Funds Available']
    df.columns = headers

    # Add new columns
    df['Future #'] = '000000'
    df['Future Name'] = 'Undefined Future'
    df['Account String'] = df.apply(lambda row: f"{row['Fund #']}-{row['Appropriation #']}-{row['Cost Center #']}-{row['Object #']}-{row['Project #']}-{row['Activity #']}-{row['Interfund #']}-{row['Future #']}", axis=1)
    df['Dept #'] = df['Cost Center #'].str[:2]

    # Adjust column order
    headers = ['Account Type', 'Dept #', 'Fund #', 'Fund Name', 
               'Appropriation #', 'Appropriation Name', 
               'Cost Center #', 'Cost Center Name', 'Object #', 
               'Object Name', 'Project #', 'Project Name', 
               'Activity #', 'Activity Name', 'Interfund #', 
               'Interfund Name', 'Future #', 'Future Name', 
               'Account String', 'Adopted Budget', 'Amended Budget', 
               'Commitment', 'Obligation', 'Actual', 'Funds Available']
    df = df[headers]

    print('cleaned data')

    # Data compression
    group_columns = headers[0:19]
    value_cols = headers[19:]
    df[value_cols] = df[value_cols].astype(float)
    df = df.groupby(group_columns, as_index=False).sum()

    print('aggregated data')

    # Filter out rows with $0.00 in specified columns
    df = df.loc[~(df[value_cols].sum(axis=1) == 0)]
    df.reset_index(drop=True, inplace=True)

    print('filtered data')

    # Save DataFrame to an Excel file with table formatting
    output_filename = 'converted_HARI.xlsx'
    output_filepath = os.path.join(OUTPUT_FOLDER, output_filename)
    with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
        print('creating table')
        df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=2)
        worksheet = writer.sheets['Sheet1']

        rows, cols = df.shape
        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            worksheet.column_dimensions[col[0].column_letter].width = max_length + 6

        table_range = f"A3:{chr(65 + cols - 1)}{rows + 3}"
        table = Table(displayName="Table1", ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
        table.tableStyleInfo = style
        worksheet.add_table(table)

        gray_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        for cell in worksheet["3:3"]:
            cell.fill = gray_fill

        for col in 'DFHJLNPR':
            worksheet.column_dimensions.group(col, col, hidden=False)

        last_row_number = rows + 4
        worksheet['S1'] = 'Total:'
        worksheet['S2'] = 'Subtotal:'
        worksheet['S1'].alignment = Alignment(horizontal='right')
        worksheet['S2'].alignment = Alignment(horizontal='right')

        for idx, col_letter in enumerate('tuvwxy', start=1):
            worksheet[f"{col_letter}{1}"] = f"=SUBTOTAL(9,{col_letter}4:{col_letter}{last_row_number - 1})"
            worksheet[f"{col_letter}{2}"] = f"=SUM({col_letter}4:{col_letter}{last_row_number - 1})"

        currency_columns = ['Adopted Budget', 'Amended Budget', 
               'Commitment', 'Obligation', 'Actual', 'Funds Available']
        
        red_text_rule = CellIsRule(operator='lessThan', formula=['0'], font=Font(color="FF0000"))

        for col_idx in range(1, cols + 1):
            column_letter = chr(64 + col_idx)
            worksheet.conditional_formatting.add(f"{column_letter}1:{column_letter}{rows + 3}", red_text_rule)

        for col in worksheet.iter_cols(min_col=1, max_col=cols, min_row=1):
            header = col[2].value
            for cell in col[:2] + col[3:]:
                cell.alignment = Alignment(horizontal='right')
                if header in currency_columns:
                    cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    print('converted')

    return output_filepath

if __name__ == '__main__':
    app.run(debug=True)