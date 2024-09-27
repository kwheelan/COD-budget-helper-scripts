from openpyxl import load_workbook
import os
from copy import copy

# ===================== Constants ===============================

# File paths
ROOT = 'Procurement'
source_file = os.listdir('Procurement/procurement_data')[0]
source_file_path = f'{ROOT}/procurement_data/{source_file}'

# Sheet names
source_sheet_name = 'Formatted Data for Detail Sheet'
destination_sheet_name = 'Non-Personnel'

# Specify columns to copy (assuming columns are labeled A, B, C, etc.)
columns_to_move = list('CDEFGHIJKLM') 
column_destinations = list('DGIKOPQRSTU')

# Start rows for copy and pasting
source_row_start = 3
destination_row_start = 19

# ================== Helper functions ===================================

# Function to copy cell and optionally its style
def copy_cell(src_cell, dest_cell, keep_dest_style = True):
    dest_cell.value = src_cell.value
    if (not keep_dest_style) and src_cell.has_style:
        dest_cell.font = copy(src_cell.font)
        dest_cell.border = copy(src_cell.border)
        dest_cell.fill = copy(src_cell.fill)
        dest_cell.number_format = src_cell.number_format
        dest_cell.protection = copy(src_cell.protection)
        dest_cell.alignment = copy(src_cell.alignment)

# function to copy given columns 
def copy_cols(columns_to_move, column_destinations, 
              source_ws, destination_ws, source_row_start, 
              destination_row_start):

    for col_ix in range(0, len(columns_to_move)):
        for row in range(0, 20): # Start with just first 10 rows source_ws.max_row + 1):

            # Get the right column letter to move from
            source_col = columns_to_move[col_ix]
            # copy cell in the right row and column
            copied_cell = source_ws[f'{source_col}{row + source_row_start}']

            # get column to paste to
            new_col = column_destinations[col_ix]
            # paste cell
            dest_cell = destination_ws[f'{new_col}{row + destination_row_start}']
            copy_cell(copied_cell, dest_cell)


# TODO: iterate through each file in detail_sheets, read its dept and division
# Translate to filters for procurement data
# filter procurement data appropriately
# copy to the file and save it

# Load workbook + worksheet from source
source_wb = load_workbook(source_file_path, data_only=False)
source_ws = source_wb[source_sheet_name]

# Load worksheets
destination_wb = load_workbook(destination_file, data_only=False)
destination_ws = destination_wb[destination_sheet_name]

# Save the destination workbook
destination_wb.save(destination_file)