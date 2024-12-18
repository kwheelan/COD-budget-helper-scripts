""" 
K. Wheelan December 2024

A script to change all summary formulas to pull from the approved budget recommendation
"""

import re
import os
from openpyxl import load_workbook, worksheet
from openpyxl.worksheet.datavalidation import DataValidation

#======================================================================
# Replacer class
#======================================================================

class Replacer:
    def __init__(self, sheet, old_col, new_col):
        self.sheet = sheet
        self.old_col = old_col
        self.new_col = new_col

    def sheet(self):
        return self.sheet
    
    def replace_function(self, formula):
        # Regular expression to match the pattern
        pattern = rf"('*{self.sheet}'*!\$){self.old_col}(\$[\d]+)?(:\$)?({self.old_col})?(\$[\d]+)?"
        
        def repl(match):
            sheet_part = match.group(1)
            first_reference_part = match.group(2) or ''
            range_separator = match.group(3) or ''
            second_reference_part = match.group(5) or ''
            
            if range_separator:
                # If range_separator is present, handle whole column and regular ranges
                return f"{sheet_part}{self.new_col}{first_reference_part}{range_separator}{self.new_col}{second_reference_part}"
            else:
                # Single cell or whole column reference
                return f"{sheet_part}{self.new_col}{first_reference_part}"

        # Replace the pattern in the text
        result = re.sub(pattern, repl, formula)
        return(result) 

#======================================================================
# Constants
#======================================================================

analyst_review_folder = 'C:/Users/katrina.wheelan/OneDrive - City of Detroit/Documents - M365-OCFO-Budget/BPA Team/FY 2026/1. Budget Development/08. Analyst Review/'
budget_director_review_folder = 'C:/Users/katrina.wheelan/OneDrive - City of Detroit/Documents - M365-OCFO-Budget/BPA Team/FY 2026/1. Budget Development/08A. Deputy Budget Director Recommend/'
SOURCE_FOLDER = budget_director_review_folder

# analyst_review_replacements = {
#     'FTE, Salary-Wage, & Benefits' : {
#         'AG' : 'AR',
#         'U': 'AO',
#         'AA': 'AP',
#         'AF': 'AQ'},
#     'Overtime & Other Personnel' : {
#         'R' : 'AC',
#         'V' : 'AD',
#         'W' : 'AE'},
#     'Revenue' : {'V' : 'Z'},
#     'Non-Personnel' : {'Y' : 'AC'}
# }

# def build_replacements(dict):
#     replacer_list = []
#     for sheet in dict.keys():
#         for old_col in dict[sheet]:
#             new_col = dict[sheet][old_col]
#             replacer = Replacer(sheet, old_col, new_col)
#             replacer_list.append(replacer)
#     return replacer_list

# REPLACEMENTS = [
#     Replacer('FTE, Salary-Wage, & Benefits', 'AG', 'AR'),
#     Replacer('FTE, Salary-Wage, & Benefits', 'U', 'AO'),
#     Replacer('FTE, Salary-Wage, & Benefits', 'AA', 'AP'),
#     Replacer('FTE, Salary-Wage, & Benefits', 'AF', 'AQ'),
#     Replacer('Overtime & Other Personnel', 'R', 'AC'),
#     Replacer('Overtime & Other Personnel', 'V', 'AD'),
#     Replacer('Overtime & Other Personnel', 'W', 'AE'),
#     Replacer('Revenue', 'V', 'Z'),
#     Replacer('Non-Personnel', 'Y', 'AC')
# ]

# POLICE_REPLACEMENTS = REPLACEMENTS.copy()
# POLICE_REPLACEMENTS[0] = Replacer('FTE, Salary-Wage, & Benefits', 'AH', 'AS')
# POLICE_REPLACEMENTS[1] = Replacer('FTE, Salary-Wage, & Benefits', 'V', 'AP')
# POLICE_REPLACEMENTS[2] = Replacer('FTE, Salary-Wage, & Benefits', 'AB', 'AQ')
# POLICE_REPLACEMENTS[3] = Replacer('FTE, Salary-Wage, & Benefits', 'AG', 'AR')

REPLACEMENTS = [
    Replacer('FTE, Salary-Wage, & Benefits', 'AR', 'BC'), #total
    Replacer('FTE, Salary-Wage, & Benefits', 'AO', 'AZ'), #FTEs
    Replacer('FTE, Salary-Wage, & Benefits', 'AP', 'AZ'), #salary/wage
    Replacer('FTE, Salary-Wage, & Benefits', 'AQ', 'BB'), #fringe
    Replacer('Overtime & Other Personnel', 'AC', 'AK'), #OT 
    Replacer('Overtime & Other Personnel', 'AD', 'AL'), # FICA
    Replacer('Overtime & Other Personnel', 'AE', 'AM'), #total
    Replacer('Revenue', 'Z', 'AD'), #total
    Replacer('Non-Personnel', 'AC', 'AG') #total
]

POLICE_REPLACEMENTS = []

#SHEETS = ['Dept Summary', 'Initiatives Summary']
SHEETS = ['Budget Director Summary', 'Initiatives Summary']

SAVE_TO = 'output/formula_replacements/'

KEY = '(Deputy Budget Director)' #'(Analyst Review)

EXCLUDE = ['Police', 'DSLP']

#======================================================================
# Helpers
#======================================================================

def replace_all(formula, police = False):
    if police:
        replacements = POLICE_REPLACEMENTS
    else: 
        replacements = REPLACEMENTS
    for replacer in replacements:
        formula = replacer.replace_function(formula)
    return formula

def find_DS(folder, keyword=KEY, exclude=EXCLUDE, verbose=False):
    # Get full file path
    folder_fp = os.path.join(SOURCE_FOLDER, folder)

    # grab list of all files in that folder
    files = os.listdir(folder_fp)

    # generate list of reviewed detail sheets
    reviewed_DS = []
    for file in files:
        if (keyword in file) and ('.xlsx' in file) and not (sum([(word in file) for word in exclude])):
            reviewed_DS.append(os.path.join(SOURCE_FOLDER, folder, file))

    # return message
    message = None
    if len(reviewed_DS) > 1:
        message = f'Multiple potential reviewed detail sheets: {reviewed_DS}'
    elif len(reviewed_DS) == 0:
        message = f'No reviewed detail sheet found in {folder}'
    if verbose and message:
        print(message)

    return reviewed_DS

def create_file_list(verbose = False):
    # all folders in analyst review section
    DS_FOLDERS = os.listdir(SOURCE_FOLDER)

    # initialize list and add viable files to it
    DS_list = []
    for folder in DS_FOLDERS:
        folder_fp = os.path.join(SOURCE_FOLDER, folder)
        if(os.path.isdir(folder_fp)):
            DS_list += find_DS(folder, keyword=KEY, exclude=EXCLUDE, verbose=verbose)

    return DS_list

# def fix_compatibility(formula):
#     # Use regex to find and remove curly braces around formulas or @
#     #formula = re.sub(r'{([^{}@]+)}', r'\1', formula)
#     #formula = formula.replace('_xlfn.', '')
#     #formula = formula.replace('_xlws.', '')
#     return formula

def edit_formulas(file, verbose=False, save_to=False):
    # check if it's the police DS
    police = 'Police' in file

    # open file 
    wb = load_workbook(file, data_only=False)

    # make replacements on both dept summary and init summary
    for sheet in SHEETS:
        ws = wb[sheet]
        
        # Iterate through all cells in the sheet
        for row in ws.iter_rows():
            for cell in row:

                if isinstance(cell.value, str) and "=" in cell.value:
                    # store old formula
                    original_formula = cell.value

                    # replace formula
                    new_formula = replace_all(original_formula, police)

                    if original_formula != new_formula:
                        # print optional message
                        if verbose:
                            print(f"Replacing {original_formula} with {new_formula} at cell {cell.coordinate}")
                        cell.value = new_formula

    # save workbook
    name = file.split('\\')[-1]
    if save_to:
        wb.save(f'{SAVE_TO}/{name}')
    else:
        wb.save(file)
    print(f'Saved {name}')

def extract_data_validations(ws):
    """
    Extract data validation settings from a worksheet.
    """
    validations = []
    for dv in ws.data_validations.dataValidation:
        # Extract the relevant properties of the data validation 
        dv_props = {
            "type": dv.type,
            "formula1": dv.formula1,
            "formula2": dv.formula2,
            "allow_blank": dv.allow_blank,
            "showDropDown": dv.showDropDown,
            "sqref": dv.sqref  # Range of cells
        }
        validations.append(dv_props)
    return validations

def get_all_data_validations(file):
    """ Save dropdowns from each sheet """
    # open file 
    wb = load_workbook(file, data_only=False)
    validations = {}
    for sheet in SHEETS.keys():
        validations[sheet] = extract_data_validations(wb[sheet])

def apply_data_validations(ws, validations):
    """
    Apply saved data validation settings to a worksheet.
    """
    for dv_props in validations:
        dv = DataValidation(
            type=dv_props["type"],
            formula1=dv_props["formula1"],
            formula2=dv_props["formula2"],
            allow_blank=dv_props["allow_blank"],
            showDropDown=dv_props["showDropDown"],
        )
        ws.add_data_validation(dv)
        dv.ranges = dv_props["sqref"]

def apply_all_data_validations(file, validations):
    """ add back all drop downs """
    # open file 
    wb = load_workbook(file, data_only=False)
    for sheet in validations.keys():
        apply_all_data_validations(wb[sheet], validations[sheet])
    wb.save(file)

#======================================================================
# Main
#======================================================================

def test():
    #text = """'FTE, Salary-Wage, & Benefits'!$AG$2 + SUM('FTE, Salary-Wage, & Benefits'!$AG$15:$AG$1000)"""
    text = """=SUMIFS('FTE, Salary-Wage, & Benefits'!$AG:$AG,'FTE, Salary-Wage, & Benefits'!$L:$L,"Baseline",'FTE, Salary-Wage, & Benefits'!$G:$G,$W7)
+SUMIFS('Overtime & Other Personnel'!$W:$W,'Overtime & Other Personnel'!$N:$N,"Baseline",'Overtime & Other Personnel'!$G:$G,$W7)
+SUMIFS('Non-Personnel'!$Y:$Y,'Non-Personnel'!$O:$O,"Baseline*",'Non-Personnel'!$G:$G,$W7)
"""
    print(replace_all(text))

def main():
    validations = get_all_data_validations(?)
    files = create_file_list()
    for file in files:
        edit_formulas(file)

if __name__ == '__main__':
    #test()
    main()