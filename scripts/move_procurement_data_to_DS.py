"""
K. Wheelan Sept 27, 2024

Iterate through each file in detail_sheets, read its dept and division
 Translate to filters for procurement data
 filter procurement data appropriately
 copy to the file and save it
"""

from openpyxl import load_workbook
import os
from copy import copy
import re
import pandas as pd

# ===================== Constants ===============================

# File paths
ROOT = 'Data'
source_file = os.listdir(f'{ROOT}/procurement_data')[0]
source_file_path = f'{ROOT}/procurement_data/{source_file}'
dest_folder = f'{ROOT}/detail_sheets'
# dest_folder = 'C:/Users/katrina.wheelan/OneDrive - City of Detroit/Documents - M365-OCFO-Budget/BPA Team/FY 2026/1. Budget Development/03. Form Development/Detail Sheets'

# Sheet names
source_sheet_name = 'Formatted Data for Detail Sheet'
destination_sheet_name = 'Non-Personnel'

# Specify columns to copy (assuming columns are labeled A, B, C, etc.)
columns_to_move = list(range(2, 13))
column_destinations = list('DGIKOPQRSTU')

# Start rows for copy and pasting
source_row_start = 3
destination_row_start = 19

# Depts
multiword_depts = [
    'City Clerk',
    'City Council',
    '36th Dist Ct'
]

# Division crosswalk
div_xwalk = pd.read_excel(f'{ROOT}/crosswalks/dept-xwalk.xlsx', sheet_name="Division")
# Create a dictionary where each key maps to a list of corresponding values
div_name_dict = div_xwalk.groupby('Nickname')['Division'].apply(list).to_dict()

# ================== Helper functions ===================================

# Function to copy cell and optionally its style
def copy_cell(src_cell, dest_cell, keep_dest_style = True):
    dest_cell.value = src_cell.value
    if (not keep_dest_style) and src_cell.has_style:
        dest_cell.font = copy(src_cell.font)
        dest_cell.border = copy(src_cell.border)
        dest_cell.fill = copy(src_cell.fill)
        dest_cell.number_format = src_cell.number_format
        dest_cell.protection = copy(src_cell.protection)
        dest_cell.alignment = copy(src_cell.alignment)

# function to copy given columns 
def copy_cols(columns_to_move, column_destinations, source_data, destination_ws, destination_row_start=1):
    """
    Copy columns from filtered data to a destination worksheet.

    Args:
        columns_to_move (list of int): Column indices from the source data to be moved (0-based).
        column_destinations (list of str): Corresponding column letters in the destination worksheet.
        source_data (list of list): Source data as a list of rows.
        destination_ws (worksheet): Destination worksheet.
        destination_row_start (int): Row number in destination worksheet to start copying (1-based index).

    Returns:
        None
    """
    for col_ix in range(len(columns_to_move)):
        for row_ix, row_data in enumerate(source_data):
            # Get the right column value from the source data
            try:
                source_col_value = row_data[columns_to_move[col_ix]]
            except:
                print(f'row_data: {row_data}')
                print(f'columns_to_move: {columns_to_move}')
                print(f'columns_to_move[col_ix]: {columns_to_move[col_ix]}')

            # get column to paste to
            new_col = column_destinations[col_ix]
            # Determine the destination cell
            dest_cell = destination_ws[f'{new_col}{row_ix + destination_row_start}']
            
            # Copy the value to the destination cell
            dest_cell.value = source_col_value

# Read data from the detail sheet file name
def extract_dept_info(filename):

    for dept in multiword_depts:
        if dept in filename:
            dept_name = dept
            # Extract department number and division
            pattern = rf"Dept (\d+)\s+{re.escape(dept)}(?:\s+(.+?))?\.xlsx"
            
            # Search for the pattern in the filename
            match = re.search(pattern, filename)
            if match:
                # Extract the department number
                dept_number = match.group(1)
                
                # Extract the department name from the known list
                dept_name = dept
                
                # Extract the division (if it exists)
                division = match.group(2) if match.group(2) else None
                
                return dept_number, dept_name, division
    
    # Fallback to single-word department names 
    pattern = r"Dept (\d+)\s+([\w-]+)(?:\s+(.+?))?\.xlsx"
    # Search for the pattern in the filename
    match = re.search(pattern, filename)
    
    if match:
        # Extract the department number + name, and the division
        dept_number = match.group(1)
         # Extract the department name from the matched pattern or from known list
        dept_name = dept_name if 'dept_name' in locals() else match.group(2)
        # Extract the division (if it exists)
        division = match.group(3) if match.group(3) else None
        return dept_number, dept_name, division
    # if no match    
    return None, None, None

def filter_data_in_excel(filename, sheet_name, header_row, filters):
    """
    Filters data in an Excel sheet based on multiple criteria.

    Args:
        filename (str): The path to the Excel file.
        sheet_name (str): The name of the sheet to filter.
        header_row (int) : The number of the header row (assume data starts in the next row)
        filters (list of tuples): List of (column_name, filter_value) pairs.

    Returns:
        list: A list of rows that match the filter criteria.
    """
    # Load the Excel workbook and the specific sheet
    workbook = load_workbook(filename, data_only=True)
    sheet = workbook[sheet_name]

    # Determine the column numbers for the filter columns (assuming the first row is headers)
    header = [cell.value for cell in sheet[header_row]]
    filter_indices = {col: header.index(col) + 1 for col, _ in filters} # +1 because openpyxl is 1-based indexing

    # Iterate through the rows and filter based on the given values
    filtered_data = []
    for row in sheet.iter_rows(min_row = (1 + header_row), values_only=True): # Skip the header row
    # for row_ix in range(header_row+1, 15):
    #     row = [cell.value for cell in sheet[row_ix]]
        match = True
        for col, value_list in filters:
            cell_content = str(row[filter_indices[col] - 1])
            # print( f'comparing col {col} to {value}: this row has value {cell_content}')
            if cell_content not in value_list: # -1 because iter_rows is 0-based indexing
                match = False
                break
        if match:
            filtered_data.append(row)

    return filtered_data

def move_data(detail_sheet):
    # Extract department information
    dept_number, _, division = extract_dept_info(detail_sheet)
    # get division name as marked in procurement data
    if division is None:
        full_div_names = ['-']
    else:
        full_div_names = div_name_dict[division]
    # filter data in source sheet based on the file's dept and div
    filters = [('Dept', [dept_number]), ('Division', full_div_names)]
    filtered_data = filter_data_in_excel(source_file_path, source_sheet_name, 
                                        source_row_start-1, filters)
    print(f'Filtering sheet for Dept: {dept_number}, Divs: {full_div_names}')
    # Load destination worksheets
    destination_wb = load_workbook(f'{dest_folder}/{detail_sheet}', data_only=False)
    destination_ws = destination_wb[destination_sheet_name]

    # copy over filtered data
    copy_cols(columns_to_move, column_destinations, filtered_data, 
            destination_ws, destination_row_start)

    # Save the destination workbook
    destination_wb.save(f'{dest_folder}/{detail_sheet}')
    print(f'saved {detail_sheet}')

#================== Program on run ============================================

def main():
    for detail_sheet in os.listdir(dest_folder):
        # only attempt on excel sheets (exlude folder, etc)
        if '.xlsx' not in detail_sheet:
            continue
        try: 
            move_data(detail_sheet)
        # Allow program to continue, even if it fails on a particular sheet (like the Matt version of Police)
        except:
            print(f'failed to process {detail_sheet}')


if __name__ == '__main__':
    main()
