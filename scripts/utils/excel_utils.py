
"""
K. Wheelan 
October 2024

Utility functions for handling Excel sheets
"""

from copy import copy
import openpyxl

# default exports
__all__ = ['copy_cell', 'copy_cols']

#----------------------- Functions -----------------------

def copy_cell(src_cell, dest_cell, keep_dest_style = True):
    """
    Function to copy cell and optionally its style """
    dest_cell.value = src_cell.value
    if (not keep_dest_style) and src_cell.has_style:
        dest_cell.font = copy(src_cell.font)
        dest_cell.border = copy(src_cell.border)
        dest_cell.fill = copy(src_cell.fill)
        dest_cell.number_format = src_cell.number_format
        dest_cell.protection = copy(src_cell.protection)
        dest_cell.alignment = copy(src_cell.alignment)


def letter_to_col(str):
    """
    Convert Excel column name (ex. AA or G) to column number (0-based)
    """
    if len(str) == 1:
        letter = str.upper()
        return ord(letter) - 65
    return 26 * (letter_to_col(str[0]) + 1) + letter_to_col(str[1:]) 

def col_to_letter(n):
    """
    Convert Excel column number (0-based) to letter (ex. A or BD)
    """
    if n < 26:
        return chr(n + 65)
    return col_to_letter( (n // 26) - 1) + col_to_letter(n % 26)

def find_next_open_row(sheet):
    """
    This function finds the first completely empty row in the given sheet.
    :param sheet: An openpyxl worksheet object.
    :return: The row number of the first completely empty row.
    """
    if type(sheet) == openpyxl.Workbook:
        raise TypeError("The object passed is a Workbook, expected a Worksheet.")
    
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_col=1, values_only=True), start=1):
        if all(cell is None for cell in row):
            return row_index
    return sheet.max_row + 1


# function to copy given columns 
def copy_cols(source_ws, destination_ws, columns_to_move, 
              column_destinations = None, destination_row_start=None, source_row_start=None):
    """
    Copy columns one worksheet to another

    Returns:
        None
    """
    # Initialize default settings
    if not column_destinations:
        column_destinations = columns_to_move

    # If no specified start row, paste into next open row
    if not destination_row_start:
        destination_row_start = find_next_open_row(destination_ws)
    if not source_row_start:
        source_row_start = 1

    # convert columns if necessary
    if (type(columns_to_move[0]) is str):
        columns_to_move = [letter_to_col(l) for l in columns_to_move]
    if (type(column_destinations[0]) is str):
        column_destinations =  [letter_to_col(l) for l in column_destinations]

    for col in range(len(columns_to_move)):
        for row in range(source_row_start, source_ws.max_row + 1):
            source_cell = source_ws.cell(row = row, column = col + 1)
            dest_cell = destination_ws.cell(row = destination_row_start + row - 1, 
                                            column = col + 1)
            copy_cell(source_cell, dest_cell)