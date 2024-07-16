import pandas as pd
import os
# from openpyxl import load_workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import re

# filter out warnings for now
import warnings
warnings.filterwarnings("ignore")

# fiscal year
FY = 25

# months of the fiscal year
MONTHS = [
    'Apr',
    'May',
    'June',
    'Placeholder for sum as of Jul 1',
    'Jul',
    'Aug',
    'Sept',
    'Oct',
    'Nov',
    'Dec',
    'Jan',
    'Feb',
    'March',
    'April',
    'May',
    'June',
    'Placeholder for sum'
]

NEW_CAL_YEAR_IX = MONTHS.index('Jan')

def fixLeadingZeroes(df, col, charLength):
    df[col] = df[col].astype(str).apply(lambda x: x.zfill(charLength) if len(x) < charLength else str(x))
    return df

def adjustCodes(df):
    # make sure that all job codes are at least 6 characters and add zeroes if not
    df = fixLeadingZeroes(df, 'Job Code', charLength=6)
    df = fixLeadingZeroes(df, 'Appropriation', charLength=5)
    df = fixLeadingZeroes(df, 'Cost Center', charLength=6)
    df = fixLeadingZeroes(df, 'Department', charLength=2)
    return df

def createAccountString(df):
    df = adjustCodes(df)
    df.loc[:, 'Account String'] = df['Department'] + df['Appropriation'] + df['Cost Center'] + df['Job Code']
    return df

def adjustColWidth(sheet):
    # Auto-adjust column widths
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                # exclude formulas but otherwise add padding to the longest cell in the col
                if ("=" not in cell.value) and (len(str(cell.value)) > max_length):
                    max_length = len(cell.value)
            except:
                pass
        # give room for table arrow
        adjusted_width = (max_length + 5)
        sheet.column_dimensions[column].width = adjusted_width

def remove_numbers_dashes(df, columns):
    for col in columns:
        df.loc[:, col] = df[col].astype(str).str.replace(r'[\d-]', '', regex=True).str.strip()
    return df

def addCols(df):
    ncols = len(df.columns)
    for i in range(len(MONTHS)):
        year = FY - (i < NEW_CAL_YEAR_IX)
        label = f"{MONTHS[i]}'{year} PAs"
        df = df.assign(**{label: pd.NA})
    # move AMJ to let of adopted FTE
    cols = df.columns.tolist()  # Get a list of all column names
    cols.insert(ncols+2, cols.pop(cols.index(f'FY{FY} Adopted FTE')))  # Pop the specified column and insert it at the beginning
    return df[cols]

def addFormulaCol(sheet, column_number, n_cols_to_sum, col_name):
    # Add a new column 'FY24 Amended FTE's' that sums the previous 13 columns
    formula_col_letter = get_column_letter(column_number)
    sheet.cell(row=1, column=column_number, value=col_name)
    
    for row in range(2, sheet.max_row + 1):
        sum_range = f"{get_column_letter(column_number-n_cols_to_sum)}{row}:{get_column_letter(column_number-1)}{row}"
        sheet[f"{formula_col_letter}{row}"] = f"=SUM({sum_range})"

def aggregateByJob(df):
    # add an ID column
    df = createAccountString(df)
    return df.groupby(['Account String']).agg({
        'Department' : 'first',
        'Fund' : 'first',
        'Fund Name' : 'first',
        'Appropriation' : 'first',
        'Appropriation Name' : 'first',
        'Cost Center' : 'first',
        'Cost Center Name' : 'first',
        'Account String' : 'first',
        'Job Code' : 'first',
        'Job Title' : 'first',
        f'FY{FY} Adopted FTE': 'sum'
    })


def addSubtotalRow(sheet, table_name, n_cols, start_col_idx):
    table = sheet.tables[table_name]
    table.showTotalRow = True  # Ensure the table's total row is shown

    # Extract the total row index correctly
    end_cell = table.ref.split(':')[1]
    total_row_idx = int(re.findall(r'\d+', end_cell)[0]) + 1  # Extract the row number and add 1
    
    # Set the "Total" label in the first column of the total row
    sheet[f"A{total_row_idx}"] = "Total"
    
    # Apply the SUBTOTAL function to the last 16 columns
    for col in range(start_col_idx, n_cols + 1):
        col_letter = get_column_letter(col)
        sheet[f"{col_letter}{total_row_idx}"] = f"=SUBTOTAL(109, {col_letter}2:{col_letter}{total_row_idx - 1})"

def format_columns_decimals(sheet, columns, n_row, decimal_places=2):
    """
    Apply number formatting to columns to have exactly two decimal places.
    """
    number_format = f"0.{''.join(['0'] * decimal_places)}"

    for col_idx in columns:
        for row in range(1, n_row + 1):
            sheet.cell(row=row, column=col_idx).number_format = number_format

def addStyling(sheet, n_rows, n_cols):
    # adjust col widths to fit content
    adjustColWidth(sheet)

    # Style the header row: bold font and white font
    header_font = Font(bold=True, color="FFFFFF")
    for cell in sheet[1]:
        cell.font = header_font

    # Style the total row: dark blue fill and white font
    total_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
    for cell in sheet[n_rows]:
        cell.font = header_font
        cell.fill = total_fill
    
    # add decimal styling to all FTE cols
    cols = range(n_cols - len(MONTHS), n_cols + 1) 
    format_columns_decimals(sheet, cols, n_rows)


# Function to add DataFrame to an openpyxl workbook and convert to table
def df_to_workbook(df, table_name, output_file):
    # trim data to delete unneccessary cols
    df = aggregateByJob(df)
    df = remove_numbers_dashes(df, ['Appropriation Name', 'Cost Center Name', 'Fund Name'])
    df = addCols(df)
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

        wb = writer.book
        sheet = wb.active

        # add a column to show starting point for positions as of July 1st
        addFormulaCol(sheet, 15, 4, f"FY{FY} Amended FTE as of 7/1/{FY-1}")

        # add a total column at the end
        addFormulaCol(sheet, sheet.max_column, 13, f"FY{FY} Amended FTE")

        # get dimensions
        n_cols = df.shape[1]
        n_rows = df.shape[0] + 1  # including the header row

        # create table with styling
        tab = Table(displayName=table_name, ref=f"A1:{get_column_letter(n_cols)}{n_rows}")
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        sheet.add_table(tab)

        # Add subtotal row for specific columns
        start_col_idx = n_cols - 17  # Adjust as necessary for your situation
        addSubtotalRow(sheet, table_name, n_cols, start_col_idx)

        # add row for substotals
        #addSubtotalRow(sheet, n_rows, n_cols)
        n_rows += 1

        # ADD SYLING
        addStyling(sheet, n_rows, n_cols)

        wb.save(output_file)  # Save the workbook


def create_workbooks(input_file, sheet_name, split_column):

    # The actual data to be split
    df = pd.read_excel(input_file, sheet_name=sheet_name, skiprows=11)

    # Get unique values from the column
    unique_values = df[split_column].unique()

    # Create a directory to save the new workbooks
    output_dir = 'Position Detail by Department'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Loop through unique values and create a new DataFrame for each
    for value in unique_values:
        subset_df = df[df[split_column] == value]

        # deal with stray values
        if value == '-':
            value = '0 - No Department Given'

        output_file = f"{output_dir}/{value}.xlsx"

        dept_num = value.split(' ')[0]

        # Write each subset to a new workbook with formatting and table conversion
        df_to_workbook(subset_df, table_name=f"Table_{dept_num}", output_file=output_file)

    print("Workbooks created successfully in the 'All Departments' directory.")

if __name__ == "__main__":
    # Define input parameters
    input_file = "Full_Position_Data.xlsx"
    sheet_name = f"FY{FY} Position Detail"
    split_column = "Department Name"
    
    create_workbooks(input_file, sheet_name, split_column)