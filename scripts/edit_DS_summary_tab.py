""" 
K. Wheelan December 2024

A script to change all summary formulas to pull from the correct detail sheet column
"""

# ======================================================================================
# SECION 1: Package imports (DO NOT EDIT)
# 
# Do not edit this section.
# ======================================================================================

import re
import os
from openpyxl import load_workbook, worksheet
from openpyxl.worksheet.datavalidation import DataValidation

# ======================================================================================
# SECION 2: Set filepaths (EDIT TO RUN)

# EDIT this section to match the file locations on your computer
# ======================================================================================

# Change the right side of this equality to be the filepath of the folder containing 
# the detail sheets to be edited. Make sure it is surrounded by quotes and 
# has the leading r (the r tells it to ignore the \s)
SOURCE_FOLDER = r'C:/Users/katrina.wheelan/OneDrive - City of Detroit/Documents - M365-OCFO-Budget/BPA Team/FY 2026/1. Budget Development/08A. Deputy Budget Director Recommend/'

# Change the right side of this equality to be the filepath of the folder to save 
# the edited detail sheets. If you set this to the same location as above, the 
# edited detail sheets will save over the older versions. Make sure the file path 
# is surrounded by quotes and has the leading r (the r tells it to ignore the \s)
SAVE_TO = r'output/formula_replacements/backups'

# ======================================================================================
# SECION 3: Set other details (EDIT TO RUN)

# EDIT this section to match what's on your detail sheets
# ======================================================================================

# Update these names to match the names of the tabs with formulas to edit.
# The tab names must be within quotes.
SHEETS = [
    'Adopted Summary', 
    'Initiatives Summary']

# Update this phrase to match text in detail sheet file name. Including this "KEY" 
# tells the program to only edit files with these words in them (and to avoid 
# touching other files in the folders).
KEY = '(Budget Director)' 

# Alter this list to include any departments to skip editing (ex. police has a 
# different column structure, so we skip it here.) Make sure any department 
# keywords are in the file names for the detail sheets to skip. Include departments
# surrounded by quotes and separated by commas.
EXCLUDE = ['Police', 'DSLP']

# How many detail sheets to alter? Default is 'All' (that aren't excluded above),
# but to test the program, you can instead put a number (without quotes), like
# N_DS = 1, which will just change the first detail sheet.
N_DS = 1 #'All'

# ======================================================================================
# SECION 4: Define columns to be swapped (EDIT TO RUN)

# The pattern is 'tab name' : { old column : new column}
# ======================================================================================

# Adjust this as needed to swap the columns in all summary formulas
REPLACEMENT_DICT = {
    'FTE, Salary-Wage, & Benefits' : {
        # Columns to be swapped in summary formulas about the FTE tab
        # Pattern is 'old_column' : 'new_column' (quotes are necessary)
        # For example, 'AG' : 'AR' will find all references to column AG
        # in the FTE tab and replace it with column AR in relevant formulas
        'AG' : 'AR',
        'U': 'AO',
        'AA': 'AP',
        'AF': 'AQ'},
    'Overtime & Other Personnel' : {
        # Columns to be swapped in summary formulas about the overtime tab
        # Pattern is 'old_column' : 'new_column' (quotes are necessary)
        'R' : 'AC',
        'V' : 'AD',
        'W' : 'AE'},
    'Revenue' : {
        # Columns to be swapped in summary formulas about the Revenue tab
        # Pattern is 'old_column' : 'new_column' (quotes are necessary)
        'V' : 'Z'},
    'Non-Personnel' : {
        # Columns to be swapped in summary formulas about the non-personnel tab
        # Pattern is 'old_column' : 'new_column' (quotes are necessary)
        'Y' : 'AC'}
}

# ======================================================================================
# SECION 5: Function and class definitions (DO NOT EDIT)

# The rest of the script tells the computer how to swap out columnn within
# the formulas. You do not need to change anything after this point. 
# ======================================================================================

def dd_ref(range) : return f"'Drop-Down Menus'!{range}"

DROPDOWN_OPTIONS = {
    'Baseline' : dd_ref('$A$7:$A$8'),
    'Baseline-Capital' : dd_ref('$A$2:$A$5'),
    'Recurring' : dd_ref('$C$2:$C$3'),
    'Employee-Type' : dd_ref('$G$2:$G$8'),
    'Position-Status' : dd_ref('$I$2:$I$4'),
    'Service' : dd_ref('$M$2:$M$11'),
    'Approval' : dd_ref('$E$2:$E$4')
}

DROPDOWNS = {
    'FTE, Salary-Wage, & Benefits' : {
        'L15:L10000' : DROPDOWN_OPTIONS['Baseline'],
        'M15:M10000' : DROPDOWN_OPTIONS['Recurring'],
        'P15:P10000' : DROPDOWN_OPTIONS['Employee-Type'],
        'Q15:Q10000' : DROPDOWN_OPTIONS['Position-Status'],
        'S15:S10000' : DROPDOWN_OPTIONS['Service'],
        'AN15:AN10000' : DROPDOWN_OPTIONS['Approval'],
        'AY15:AY10000' : DROPDOWN_OPTIONS['Approval']
    },
    'Overtime & Other Personnel' : {
        'N15:N10000' : DROPDOWN_OPTIONS['Baseline'],
        'O15:O10000' : DROPDOWN_OPTIONS['Recurring'],
        'Q15:Q10000' : DROPDOWN_OPTIONS['Service'],
        'AB15:AB10000' : DROPDOWN_OPTIONS['Approval'],
        'AJ15:AJ10000' : DROPDOWN_OPTIONS['Approval']
    },
    'Non-Personnel' : {
        'O19:O10000' : DROPDOWN_OPTIONS['Baseline-Capital'],
        'P19:P10000' : DROPDOWN_OPTIONS['Recurring'],
        'X19:X10000' : DROPDOWN_OPTIONS['Service'],
        'AB19:AB10000' : DROPDOWN_OPTIONS['Approval'],
        'AF19:AF10000' : DROPDOWN_OPTIONS['Approval']
    },
    'Revenue' : {
        'P15:P10000' : DROPDOWN_OPTIONS['Baseline-Capital'],
        'Q15:Q10000' : DROPDOWN_OPTIONS['Recurring'],
        'Y15:Y10000' : DROPDOWN_OPTIONS['Approval'],
        'AC15:AC10000' : DROPDOWN_OPTIONS['Approval']
    },
}

#======================================================================
# Replacer class
#======================================================================

class Replacer:
    def __init__(self, sheet, old_col, new_col):
        self.sheet = sheet
        self.old_col = old_col
        self.new_col = new_col

    def sheet(self):
        return self.sheet
    
    def replace_function(self, formula):
        # Regular expression to match the pattern
        pattern = rf"('*{self.sheet}'*!\$){self.old_col}(\$[\d]+)?(:\$)?({self.old_col})?(\$[\d]+)?"
        
        def repl(match):
            sheet_part = match.group(1)
            first_reference_part = match.group(2) or ''
            range_separator = match.group(3) or ''
            second_reference_part = match.group(5) or ''
            
            if range_separator:
                # If range_separator is present, handle whole column and regular ranges
                return f"{sheet_part}{self.new_col}{first_reference_part}{range_separator}{self.new_col}{second_reference_part}"
            else:
                # Single cell or whole column reference
                return f"{sheet_part}{self.new_col}{first_reference_part}"

        # Replace the pattern in the text
        result = re.sub(pattern, repl, formula)
        return(result) 


def build_replacements(dict):
    replacer_list = []
    for sheet in dict.keys():
        for old_col in dict[sheet]:
            new_col = dict[sheet][old_col]
            replacer = Replacer(sheet, old_col, new_col)
            replacer_list.append(replacer)
    return replacer_list


POLICE_REPLACEMENTS = []
REPLACEMENTS = build_replacements(REPLACEMENT_DICT)

#======================================================================
# Helpers
#======================================================================

def replace_all(formula, police = False):
    if police:
        replacements = POLICE_REPLACEMENTS
    else: 
        replacements = REPLACEMENTS
    for replacer in replacements:
        formula = replacer.replace_function(formula)
    return formula

def find_DS(folder, keyword=KEY, exclude=EXCLUDE, verbose=False):
    # Get full file path
    folder_fp = os.path.join(SOURCE_FOLDER, folder)

    # grab list of all files in that folder
    files = os.listdir(folder_fp)

    # generate list of reviewed detail sheets
    reviewed_DS = []
    for file in files:
        if (keyword in file) and ('.xlsx' in file) and not (sum([(word in file) for word in exclude])):
            reviewed_DS.append(os.path.join(SOURCE_FOLDER, folder, file))

    # return message
    message = None
    if len(reviewed_DS) > 1:
        message = f'Multiple potential detail sheets in current folder: {reviewed_DS}'
    elif len(reviewed_DS) == 0:
        message = f'No detail sheet found in {folder}'
    if verbose and message:
        print(message)

    return reviewed_DS

def create_file_list(verbose = False):
    # all folders in analyst review section
    DS_FOLDERS = os.listdir(SOURCE_FOLDER)

    # initialize list and add viable files to it
    DS_list = []
    for folder in DS_FOLDERS:
        folder_fp = os.path.join(SOURCE_FOLDER, folder)
        if(os.path.isdir(folder_fp)):
            DS_list += find_DS(folder, keyword=KEY, exclude=EXCLUDE, verbose=verbose)
    return DS_list

def edit_formulas(file, verbose=False, save_to=False, backup=False):
    # check if it's the police DS
    police = 'Police' in file

    # open file 
    wb = load_workbook(file, data_only=False)

    name = file.split('\\')[-1]
    if(backup):
        wb.save(f'{SAVE_TO}/{name}')

    # make replacements on both dept summary and init summary
    for sheet in SHEETS:
        ws = wb[sheet]
        
        # Iterate through all cells in the sheet
        for row in ws.iter_rows():
            for cell in row:

                if isinstance(cell.value, str) and "=" in cell.value:
                    # store old formula
                    original_formula = cell.value

                    # replace formula
                    new_formula = replace_all(original_formula, police)

                    if original_formula != new_formula:
                        # print optional message
                        if verbose:
                            print(f"Replacing {original_formula} with {new_formula} at cell {cell.coordinate}")
                        cell.value = new_formula

    # Add back drop down menus
    add_dropdowns(wb)

    # save workbook
    name = file.split('\\')[-1]
    if save_to:
        wb.save(f'{SAVE_TO}/{name}')
    else:
        wb.save(file)
    print(f'Edited and saved {name}')

def add_dropdowns(wb):
    for sheet in DROPDOWNS.keys():
        ws = wb[sheet]

        # for each column
        for cell_range in DROPDOWNS[sheet].keys():
                
            # build data validation
            dv = DataValidation(
                type="list",
                formula1=f'={DROPDOWNS[sheet][cell_range]}',
                showDropDown=False
            )

            # Add the data validation to a specific cell, e.g., A1
            ws.add_data_validation(dv)
            dv.add(cell_range)


#======================================================================
# Main
#======================================================================

def test():
    text = """'FTE, Salary-Wage, & Benefits'!$AG$2 + SUM('FTE, Salary-Wage, & Benefits'!$AG$15:$AG$1000)"""
    print(replace_all(text))

def main():
    files = create_file_list()
    if type(N_DS) is int:
        files = files[0:N_DS]
    for file in files:
        edit_formulas(file, backup=SAVE_TO)

if __name__ == '__main__':
    #test()
    main()